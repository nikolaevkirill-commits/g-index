#!/usr/bin/env python3
"""
run_forecast.py — ЄДИНИЙ автоматичний прогноз G-Index.
Version: v1.0 (2026-06-14)

Один запуск робить ВСЕ:
  1. Тягне свіжий Kp з NOAA → future_kp.json (через fetch_kp_v2)
  2. Читає Excel теги Таніти (ДАНІ_ЩОДЕННІ, col N) + calendar enrichment
  3. Рахує score_day_v19 (з реальним Kp, поза freeze)
  4. Накладає expert_overrides (де є PDF бал — він пріоритет)
  5. Виводить прогноз: CSV + MD + JSON для дашборду

ЗАПУСК:
    python run_forecast.py                          # 14 днів від сьогодні
    python run_forecast.py --from 2026-06-29 --days 14
    python run_forecast.py --no-fetch               # без оновлення Kp (швидко)

ВИХІД (у поточну папку):
    forecast_<from>_<to>.csv   — таблиця
    forecast_<from>_<to>.md    — читабельний бюлетень
    forecast_<from>_<to>.json  — для дашборду / downstream

НЕ змінює: engine_scores.json, forecast_engine_v18_5.py (V3 freeze).
"""
import argparse
import csv
import json
import subprocess
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

# Накшатра текст→номер (з panchanga_sign_priors.json)
# Завантажується при старті; fallback = вбудований словник
def _load_nak_map(workdir: Path) -> dict:
    for d in (workdir, Path(__file__).resolve().parent):
        fp = d / 'panchanga_sign_priors.json'
        if fp.exists():
            try:
                return json.loads(fp.read_text(encoding='utf-8')).get('nakshatra_text_to_num', {})
            except Exception:
                pass
    return {
        'Ashwini':1,'Bharani':2,'Krittika':3,'Rohini':4,'Mrigashira':5,
        'Ardra':6,'Punarvasu':7,'Pushya':8,'Ashlesha':9,'Magha':10,
        'P.Phalguni':11,'U.Phalguni':12,'Hasta':13,'Chitra':14,'Swati':15,
        'Vishakha':16,'Anuradha':17,'Jyeshtha':18,'Mula':19,'P.Ashadha':20,
        'U.Ashadha':21,'Shravana':22,'Dhanishta':23,'Shatabhisha':24,
        'P.Bhadrapada':25,'U.Bhadrapada':26,'Revati':27,
    }

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
sys.path.insert(0, '/mnt/project')

# ── Verdict mapping ──────────────────────────────────────────────────
VERDICT = {
    -3: ('🔴', 'Особливо несприятливий день'),
    -2: ('🟠', 'Несприятливий день'),
    -1: ('🟡', 'Помірно несприятливий день'),
     0: ('⚪', 'Нейтральний день'),
     1: ('🟢', 'Помірно сприятливий день'),
     2: ('🟢', 'Сприятливий день'),
     3: ('💚', 'Особливо сприятливий день'),
}
DOW_UK = ['Понеділок', 'Вівторок', 'Середа', 'Четвер', "П'ятниця", 'Субота', 'Неділя']


def log(msg):
    print(f"  {msg}", flush=True)


def step_fetch_kp(workdir: Path, skip: bool):
    """Крок 1: оновити future_kp.json."""
    out = workdir / 'future_kp.json'
    if skip:
        log(f"[1/5] Kp: пропущено (--no-fetch). Використовую наявний {out.name}")
        return
    fetch = None
    for name in ('fetch_kp_v2.py', 'fetch_kp_v2__1_.py'):
        if (workdir / name).exists():
            fetch = name
            break
    if not fetch:
        log("[1/5] Kp: fetch_kp_v2.py не знайдено — пропускаю (наявний future_kp.json)")
        return
    log(f"[1/5] Kp: {fetch} --export-future …")
    try:
        r = subprocess.run([sys.executable, fetch, '--export-future', str(out)],
                           cwd=workdir, timeout=120)
        if r.returncode == 0:
            log(f"      ✅ {out.name} оновлено")
        else:
            log(f"      ⚠ rc={r.returncode} — використаю наявний future_kp.json")
    except Exception as e:
        log(f"      ⚠ {e} — використаю наявний future_kp.json")


def load_json(path: Path, default):
    try:
        return json.loads(path.read_text(encoding='utf-8'))
    except Exception:
        return default


def step_load_inputs(workdir: Path):
    """Крок 2: завантажити всі джерела."""
    import openpyxl

    # future_kp
    fkp_raw = load_json(workdir / 'future_kp.json', {})
    fkp = dict(fkp_raw.get('kp', {}))
    log(f"[2/5] future_kp: {len(fkp)} днів")

    # GFZ архів реального Kp (історичні дати, замінює synthetic). Не перетирає future_kp.
    gfz_raw = load_json(workdir / 'gfz_kp_archive.json', {})
    gfz = gfz_raw.get('kp', {})
    added = 0
    for ds, rec in gfz.items():
        if ds not in fkp:
            fkp[ds] = rec if isinstance(rec, dict) else {'kp': rec, 'source': 'gfz_archive'}
            added += 1
    if added:
        log(f"      gfz_archive: +{added} історичних днів реального Kp")

    # expert_overrides
    ov_raw = load_json(workdir / 'expert_overrides_v3.json', {})
    ov = {r['date']: r['expert_eng'] for r in ov_raw.get('overrides', [])}
    log(f"      overrides: {len(ov)} записів")

    # Excel теги
    xl = {}
    xlsx = None
    for name in ('prognoz_2025_2026_4_FIXED.xlsx', 'prognoz_2025_2026_4.xlsx'):
        if (workdir / name).exists():
            xlsx = workdir / name
            break
    if xlsx:
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb['ДАНІ_ЩОДЕННІ'] if 'ДАНІ_ЩОДЕННІ' in wb.sheetnames else wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            c = row[0]
            ds = None
            if isinstance(c, datetime):
                ds = c.date().isoformat()
            elif isinstance(c, date):
                ds = c.isoformat()
            if ds:
                xl[ds] = {
                    'tag': str(row[13] or '').strip(),   # col N
                    'kp':  row[9],                        # col J
                    'sn':  row[11] or 0,                  # col L
                }
        log(f"      Excel: {len(xl)} днів ({xlsx.name})")
    else:
        log("      ⚠ Excel не знайдено — теги порожні")

    # annual (tithi_n для P-v19-5 панчанга-пріору)
    an_raw = load_json(workdir / 'annual_2026_27.json', {})
    an = {d['date']: d for d in an_raw.get('days', [])}
    log(f"      annual: {len(an)} днів (tithi для панчанга-пріору)")

    return fkp, ov, xl, an


def step_compute(start: date, days: int, fkp, ov, xl, an):
    """Крок 3+4: рахуємо score + накладаємо overrides."""
    try:
        from score_engine_v19_preview import score_day_v19
        engine_name = 'v19_preview'
    except ImportError:
        from forecast_engine_v18_5 import score_day as score_day_v19
        engine_name = 'v18.5 (fallback)'
    log(f"[3/5] Engine: {engine_name}")

    rows = []
    for i in range(days):
        d = start + timedelta(i)
        ds = d.isoformat()
        x = xl.get(ds, {})
        tag = x.get('tag', '')
        # Kp пріоритет: future_kp → Excel → 2.0
        kp = None
        kp_src = ''
        if ds in fkp and isinstance(fkp[ds], dict):
            kp = fkp[ds].get('kp')
            kp_src = fkp[ds].get('source', '27do')
        if kp is None and x.get('kp') is not None:
            kp = float(x['kp'])
            kp_src = 'excel'
        if kp is None:
            kp = 2.0
            kp_src = 'default'
        sn = x.get('sn', 0)

        tithi_n = an.get(ds, {}).get('tithi_n')
        # nakshatra: annual має текст → конвертуємо через panchanga_sign_priors
        nak_text = an.get(ds, {}).get('nakshatra', '')
        nakshatra_n = _NAK_MAP.get(nak_text) if nak_text else None
        eng = score_day_v19(tag, kp, sn=sn, date_str=ds, tithi_n=tithi_n, nakshatra_n=nakshatra_n)
        exp = ov.get(ds)
        final = exp if exp is not None else eng
        source = 'PDF✓' if exp is not None else 'engine~'

        emoji, verdict = VERDICT.get(final, ('⚪', '?'))
        rows.append({
            'date': ds,
            'dow': DOW_UK[d.weekday()],
            'tag': tag,
            'kp': kp,
            'kp_source': kp_src,
            'engine': eng,
            'expert': exp,
            'final': final,
            'source': source,
            'emoji': emoji,
            'verdict': verdict,
        })
    log(f"[4/5] Overrides: {sum(1 for r in rows if r['expert'] is not None)}/{days} днів з PDF")
    return rows, engine_name


def step_output(rows, start: date, end: date, workdir: Path, engine_name: str):
    """Крок 5: записати CSV + MD + JSON."""
    tag = f"{start.isoformat()}_{end.isoformat()}"

    # CSV
    csv_path = workdir / f'forecast_{tag}.csv'
    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f, delimiter=';')
        w.writerow(['Дата', 'День', 'Бал', 'Вердикт', 'Kp', 'Kp_джерело', 'Engine', 'Expert', 'Джерело', 'Тег'])
        for r in rows:
            w.writerow([r['date'], r['dow'], r['final'], r['verdict'],
                       r['kp'], r['kp_source'], r['engine'], r['expert'], r['source'], r['tag']])

    # MD
    md_path = workdir / f'forecast_{tag}.md'
    lines = [f"# G-Index Прогноз {start.isoformat()} — {end.isoformat()}",
             f"\n> Engine: {engine_name} · згенеровано {date.today().isoformat()}\n"]
    for r in rows:
        ks = f"Kp={r['kp']:.1f} ({r['kp_source']})"
        lines.append(f"**{r['date']} ({r['dow']})** {r['emoji']} {r['verdict']} "
                     f"`G={r['final']:+d}` · {ks} · [{r['source']}]")
    md_path.write_text('\n\n'.join(lines), encoding='utf-8')

    # JSON
    json_path = workdir / f'forecast_{tag}.json'
    payload = {
        'version': 'run_forecast_v1',
        'generated': date.today().isoformat(),
        'engine': engine_name,
        'range': {'from': start.isoformat(), 'to': end.isoformat()},
        'days': [{k: r[k] for k in ('date', 'dow', 'final', 'verdict', 'kp',
                                     'kp_source', 'engine', 'expert', 'source', 'tag')}
                 for r in rows],
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')

    log(f"[5/5] Вихід:")
    log(f"      {csv_path.name}")
    log(f"      {md_path.name}")
    log(f"      {json_path.name}")
    return csv_path, md_path, json_path


def main():
    ap = argparse.ArgumentParser(description='G-Index автоматичний прогноз')
    ap.add_argument('--from', dest='start', default=None, help='Старт YYYY-MM-DD (default: сьогодні)')
    ap.add_argument('--days', type=int, default=14)
    ap.add_argument('--workdir', default=str(HERE), help='Папка проєкту')
    ap.add_argument('--no-fetch', action='store_true', help='Не оновлювати Kp')
    args = ap.parse_args()

    workdir = Path(args.workdir)
    start = (datetime.strptime(args.start, '%Y-%m-%d').date()
             if args.start else date.today())
    end = start + timedelta(args.days - 1)

    # Завантажуємо NAK_MAP один раз
    global _NAK_MAP
    _NAK_MAP = _load_nak_map(workdir)

    print("=" * 60)
    print(f"  G-INDEX АВТОПРОГНОЗ {start} → {end} ({args.days} днів)")
    print("=" * 60)

    step_fetch_kp(workdir, args.no_fetch)
    fkp, ov, xl, an = step_load_inputs(workdir)
    rows, engine_name = step_compute(start, args.days, fkp, ov, xl, an)
    step_output(rows, start, end, workdir, engine_name)

    print("\n" + "=" * 60)
    print("  ГОТОВО. Прогноз нижче:")
    print("=" * 60)
    for r in rows:
        print(f"  {r['date']} {r['dow']:<10} {r['emoji']} {r['verdict']:<28} "
              f"G={r['final']:+d}  Kp={r['kp']:.1f}  [{r['source']}]")


if __name__ == '__main__':
    main()
