#!/usr/bin/env python3
"""
generate_forecast_pdf.py — Generate weekly forecast PDF from Excel data.

Reads `prognoz_2025_2026_4.xlsx` (sheet ДАНІ_ЩОДЕННІ) and produces a
multi-page PDF bulletin in the format compatible with parse_forecast_pdf.py
ground-truth extraction.

Output structure: 7-column weekly tables with verdict text + recommendations,
using DejaVu Serif font for proper Cyrillic rendering.

USAGE
-----
    # Generate forecast for date range
    python generate_forecast_pdf.py --start 2026-05-04 --end 2026-08-01 \
                                    --out forecast_v3_freeze.pdf

    # Append to existing PDF
    python generate_forecast_pdf.py --start 2026-08-04 --end 2026-08-31 \
                                    --append-to existing.pdf \
                                    --out merged.pdf

    # Specific weeks only
    python generate_forecast_pdf.py --weeks 2026-W18,2026-W19 --out wk.pdf
"""

import argparse
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install: pip install openpyxl",
          file=sys.stderr)
    sys.exit(2)

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
except ImportError:
    print("ERROR: reportlab required. Install: pip install reportlab",
          file=sys.stderr)
    sys.exit(2)

try:
    from pypdf import PdfReader, PdfWriter
    HAS_PYPDF = True
except ImportError:
    HAS_PYPDF = False


# ──────────────────────────────────────────────────────────────────────
# Font registration (DejaVu Serif for Cyrillic)
# ──────────────────────────────────────────────────────────────────────

DEJAVU_PATHS = [
    '/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf',
    '/usr/share/fonts/dejavu/DejaVuSerif.ttf',
    'C:/Windows/Fonts/DejaVuSerif.ttf',
    str(Path(__file__).parent / 'DejaVuSerif.ttf'),
]

DEJAVU_BOLD_PATHS = [
    '/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf',
    '/usr/share/fonts/dejavu/DejaVuSerif-Bold.ttf',
    'C:/Windows/Fonts/DejaVuSerif-Bold.ttf',
    str(Path(__file__).parent / 'DejaVuSerif-Bold.ttf'),
]

FONT_NAME = 'DejaVuSerif'
FONT_NAME_BOLD = 'DejaVuSerif-Bold'

def _register_fonts():
    """Register DejaVu Serif. Falls back to Helvetica if not found."""
    global FONT_NAME, FONT_NAME_BOLD
    found = False
    for p in DEJAVU_PATHS:
        if Path(p).exists():
            try:
                pdfmetrics.registerFont(TTFont(FONT_NAME, p))
                found = True
                break
            except Exception:
                continue
    if not found:
        print("[WARN] DejaVu Serif not found; Cyrillic may render incorrectly. "
              "Install: apt-get install fonts-dejavu", file=sys.stderr)
        FONT_NAME = 'Helvetica'

    bold_found = False
    for p in DEJAVU_BOLD_PATHS:
        if Path(p).exists():
            try:
                pdfmetrics.registerFont(TTFont(FONT_NAME_BOLD, p))
                bold_found = True
                break
            except Exception:
                continue
    if not bold_found:
        FONT_NAME_BOLD = 'Helvetica-Bold'


# ──────────────────────────────────────────────────────────────────────
# Verdict mapping (eng score → Ukrainian phrase + cell color)
# ──────────────────────────────────────────────────────────────────────

VERDICT_MAP = {
    -3: ('Особливо несприятливий день', colors.HexColor('#d32f2f')),
    -2: ('Несприятливий день',          colors.HexColor('#f57c00')),
    -1: ('Помірно несприятливий день',  colors.HexColor('#fbc02d')),
     0: ('Нейтральний день',            colors.HexColor('#9e9e9e')),
     1: ('Помірно сприятливий день',    colors.HexColor('#aed581')),
     2: ('Сприятливий день',            colors.HexColor('#7cb342')),
     3: ('Особливо сприятливий день',   colors.HexColor('#388e3c')),
}

WEEKDAYS_UK = ['Понеділок', 'Вівторок', 'Середа', 'Четвер',
               'П\'ятниця', 'Субота', 'Неділя']


# ──────────────────────────────────────────────────────────────────────
# Excel reader
# ──────────────────────────────────────────────────────────────────────

def load_excel_forecast(xlsx_path: Path, start: date, end: date):
    """Read forecast rows from ДАНІ_ЩОДЕННІ in [start, end].

    Returns dict iso_date → {tag, kp, eng, prognoz_text}.
    Columns: A=date, J=Kp, N=tag, P=weekly forecast text.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if 'ДАНІ_ЩОДЕННІ' not in wb.sheetnames:
        raise KeyError(f"Sheet 'ДАНІ_ЩОДЕННІ' not in {xlsx_path}")
    ws = wb['ДАНІ_ЩОДЕННІ']
    out = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, datetime):
            v = v.date()
        if not isinstance(v, date):
            continue
        if v < start or v > end:
            continue
        out[v.isoformat()] = {
            'tag':     str(ws.cell(row=r, column=14).value or '').strip(),
            'kp':      ws.cell(row=r, column=10).value,
            'sn':      ws.cell(row=r, column=12).value,
            'prognoz': str(ws.cell(row=r, column=16).value or '').strip(),
        }
    return out


# ──────────────────────────────────────────────────────────────────────
# Engine integration (compute eng score if not in xlsx)
# ──────────────────────────────────────────────────────────────────────

def compute_eng_for_day(tag: str, kp, engine_module=None):
    """Compute engine score using v18.5 if available, else v17."""
    try:
        kp_f = float(kp) if kp is not None else 2.0
    except (TypeError, ValueError):
        kp_f = 2.0
    if engine_module is None:
        for modname in ('score_engine_v19_preview', 'forecast_engine_v18_5', 'forecast_engine_v17_0'):
            try:
                engine_module = __import__(modname)
                break
            except ImportError:
                continue
    if engine_module is None:
        return 0
    try:
        return engine_module.score_day(tag, kp_f)
    except Exception:
        return 0


# ──────────────────────────────────────────────────────────────────────
# Week grouping
# ──────────────────────────────────────────────────────────────────────

def iter_weeks(start: date, end: date):
    """Yield (mon, sun) tuples covering Mon-Sun weeks intersecting [start,end]."""
    # Move start back to Monday
    mon = start - timedelta(days=start.weekday())
    while mon <= end:
        sun = mon + timedelta(days=6)
        yield mon, sun
        mon = sun + timedelta(days=1)


# ──────────────────────────────────────────────────────────────────────
# PDF rendering
# ──────────────────────────────────────────────────────────────────────

def build_week_table(week_data, mon: date, sun: date, styles):
    """Build a 7-column Table for one week. week_data = dict iso → row dict."""
    # Row 1: dates (DD.MM.YYYY)
    row_dates = []
    row_weekdays = []
    row_verdicts = []
    row_recs = []
    cell_colors = []

    for i in range(7):
        d = mon + timedelta(days=i)
        iso = d.isoformat()
        rec = week_data.get(iso, {})
        eng = rec.get('eng', 0) if rec else 0
        verdict_text, color = VERDICT_MAP.get(eng, VERDICT_MAP[0])

        row_dates.append(d.strftime('%d.%m.%Y'))
        row_weekdays.append(WEEKDAYS_UK[i])
        row_verdicts.append(Paragraph(verdict_text, styles['cell']))
        row_recs.append(Paragraph(rec.get('prognoz', '') or '—', styles['rec']))
        cell_colors.append(color)

    data = [row_dates, row_weekdays, row_verdicts, row_recs]
    col_widths = [(28 * cm) / 7] * 7

    tbl = Table(data, colWidths=col_widths,
                rowHeights=[0.8 * cm, 0.7 * cm, 1.5 * cm, None])

    style = TableStyle([
        ('FONT', (0, 0), (-1, -1), FONT_NAME, 8),
        ('ALIGN', (0, 0), (-1, 1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONT', (0, 0), (-1, 0), FONT_NAME_BOLD, 9),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ])
    # Color verdict row by score
    for i, c in enumerate(cell_colors):
        style.add('BACKGROUND', (i, 2), (i, 2), c)
        style.add('TEXTCOLOR', (i, 2), (i, 2), colors.white)
    tbl.setStyle(style)
    return tbl


def build_pdf(forecast_data: dict, out_path: Path, title: str = "G-Index Forecast"):
    """Render PDF from forecast_data (dict iso → row)."""
    _register_fonts()

    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=landscape(A4),
        rightMargin=1 * cm, leftMargin=1 * cm,
        topMargin=1 * cm, bottomMargin=1 * cm,
        title=title, author="G-Index v18.5",
    )

    base = getSampleStyleSheet()
    styles = {
        'h1': ParagraphStyle('h1', parent=base['Title'],
                             fontName=FONT_NAME_BOLD, fontSize=14,
                             alignment=TA_CENTER, spaceAfter=6),
        'h2': ParagraphStyle('h2', parent=base['Heading2'],
                             fontName=FONT_NAME_BOLD, fontSize=11,
                             alignment=TA_CENTER, spaceAfter=4),
        'cell': ParagraphStyle('cell', parent=base['BodyText'],
                               fontName=FONT_NAME, fontSize=8,
                               alignment=TA_CENTER, leading=10,
                               textColor=colors.white),
        'rec': ParagraphStyle('rec', parent=base['BodyText'],
                              fontName=FONT_NAME, fontSize=7,
                              alignment=TA_LEFT, leading=9),
        'foot': ParagraphStyle('foot', parent=base['BodyText'],
                               fontName=FONT_NAME, fontSize=7,
                               alignment=TA_LEFT, textColor=colors.grey),
    }

    if not forecast_data:
        raise ValueError("No forecast data in date range")

    dates_sorted = sorted(forecast_data.keys())
    start = date.fromisoformat(dates_sorted[0])
    end = date.fromisoformat(dates_sorted[-1])

    story = []
    story.append(Paragraph(title, styles['h1']))
    story.append(Paragraph(
        f"Період: {start.isoformat()} → {end.isoformat()} "
        f"(Engine v18.5, R&amp;D / Advisory)",
        styles['h2']))
    story.append(Spacer(1, 4 * mm))

    for week_no, (mon, sun) in enumerate(iter_weeks(start, end)):
        if week_no > 0 and week_no % 2 == 0:
            story.append(PageBreak())
        # Filter forecast_data to this week
        week_data = {}
        for i in range(7):
            d = mon + timedelta(days=i)
            iso = d.isoformat()
            if iso in forecast_data:
                week_data[iso] = forecast_data[iso]
        if not week_data:
            continue
        story.append(build_week_table(week_data, mon, sun, styles))
        story.append(Spacer(1, 5 * mm))

    # Footer disclaimer
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph(
        "Прогноз здійснено з урахуванням впливу космофізичних факторів: "
        "сонячні бурі (прогнозований К-індекс), вплив Місяця, астрологічний "
        "вплив. Інформація носить рекомендаційний характер. "
        "Engine v18.5 — internal validation, prospective freeze active.",
        styles['foot']))

    doc.build(story)
    return out_path


# ──────────────────────────────────────────────────────────────────────
# PDF append utility
# ──────────────────────────────────────────────────────────────────────

def append_pdfs(base_path: Path, append_path: Path, out_path: Path):
    """Concatenate two PDFs into out_path."""
    if not HAS_PYPDF:
        raise ImportError("pypdf required for --append-to. pip install pypdf")
    writer = PdfWriter()
    for p in (base_path, append_path):
        reader = PdfReader(str(p))
        for page in reader.pages:
            writer.add_page(page)
    with open(out_path, 'wb') as f:
        writer.write(f)


# ──────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Generate weekly forecast PDF from Excel data"
    )
    ap.add_argument('--xlsx',
                    default='/mnt/project/prognoz_2025_2026_4.xlsx',
                    help='Source Excel file')
    ap.add_argument('--start', help='Start date YYYY-MM-DD')
    ap.add_argument('--end', help='End date YYYY-MM-DD')
    ap.add_argument('--out', required=True, help='Output PDF path')
    ap.add_argument('--append-to', help='Append output to existing PDF')
    ap.add_argument('--title', default='G-Index Weekly Forecast',
                    help='PDF title')
    ap.add_argument('--no-engine', action='store_true',
                    help="Don't compute eng scores (use Excel column only)")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: Excel not found: {xlsx_path}", file=sys.stderr)
        return 2

    # Date range defaults: today → today + 90 days
    today = date.today()
    start = date.fromisoformat(args.start) if args.start else today
    end = date.fromisoformat(args.end) if args.end else (today + timedelta(days=90))

    print(f"Loading: {xlsx_path}")
    print(f"  range: {start} → {end}")
    rows = load_excel_forecast(xlsx_path, start, end)
    print(f"  rows:  {len(rows)}")

    if not rows:
        print("ERROR: no rows in date range", file=sys.stderr)
        return 1

    # Compute eng scores
    if not args.no_engine:
        eng_mod = None
        for modname in ('score_engine_v19_preview', 'forecast_engine_v18_5', 'forecast_engine_v17_0'):
            try:
                eng_mod = __import__(modname)
                print(f"  engine: {modname}")
                break
            except ImportError:
                continue
        for iso, rec in rows.items():
            rec['eng'] = compute_eng_for_day(rec['tag'], rec['kp'], eng_mod)

    out_path = Path(args.out)
    if args.append_to:
        tmp = out_path.with_suffix('.tmp.pdf')
        build_pdf(rows, tmp, title=args.title)
        append_pdfs(Path(args.append_to), tmp, out_path)
        tmp.unlink()
        print(f"Appended: {args.append_to} + new → {out_path}")
    else:
        build_pdf(rows, out_path, title=args.title)
        print(f"Saved: {out_path}")
    return 0


if __name__ == '__main__':
    sys.exit(main())
