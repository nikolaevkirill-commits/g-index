"""
update_excel.py — Оновлює col P (ПРОГНОЗ тижневий) в prognoz_2025_2026.xlsx
Використовує forecast_engine.py і fetch_kp.py

Використання:
  python update_excel.py                    # оновити всі порожні P
  python update_excel.py --all              # перезаписати всі P (включно з існуючими)
  python update_excel.py --date 2026-03-09  # тільки один день
  python update_excel.py --from 2026-03-01 --to 2026-03-31  # діапазон
"""

import sys, os, shutil, argparse
from datetime import date, datetime, timedelta
from pathlib import Path

# ── Імпорт модулів ────────────────────────────────────────────────────
script_dir = Path(__file__).parent
sys.path.insert(0, str(script_dir))

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    print("❌ Встанови openpyxl: pip install openpyxl")
    sys.exit(1)

from forecast_engine_v14_6 import format_day, score_day, label
from fetch_kp_v2 import fetch_all, storm_level

# ── Колонки (1-based для openpyxl) ───────────────────────────────────
COL_DATE   = 1   # A
COL_KP     = 10  # J
COL_STORM  = 11  # K
COL_SN     = 12  # L
COL_DST    = 13  # M — Dst index (нТл) NEW v14.3
COL_JY     = 14  # N — Джйотіш тег
COL_SCORE  = 15  # O — Оцінка дня
COL_P      = 16  # P — ПРОГНОЗ тижневий

# ── Кольори по оцінці ─────────────────────────────────────────────────
SCORE_COLORS = {
    4:  "00B050",   # ОС найкращий
    3:  "70AD47",   # ОС
    2:  "E2EFDA",   # Сприятливий
    1:  "FFEB9C",   # Помірно сприятливий
    0:  "DDEBF7",   # Нейтральний
   -1:  "FCE4D6",   # Помірно несприятливий
   -2:  "FFC7CE",   # Несприятливий
   -3:  "FF7575",   # Особливо несприятливий
}

SCORE_LABELS_SHORT = {
    4: "⭐Найкращий", 3: "✅Особливо сприятл.", 2: "✅Сприятливий",
    1: "🟡Помірно сприятл.", 0: "⚪Нейтральний",
   -1: "🔶Помірно несприятл.", -2: "🔴Несприятливий", -3: "🔴Особливо несприятл."
}

def parse_date(cell_val):
    if cell_val is None: return None
    if isinstance(cell_val, (date, datetime)):
        return cell_val.date() if isinstance(cell_val, datetime) else cell_val
    try:
        return date.fromisoformat(str(cell_val)[:10])
    except:
        return None

def detect_eclipse_context(ws, row_idx, window=7):
    """Перевіряємо чи є затемнення в ±window рядків"""
    for delta in range(-window, window+1):
        r = row_idx + delta
        if r < 2: continue
        jy = ws.cell(r, COL_JY).value or ''
        if 'затемнення' in jy.lower():
            return True
    return False

def run_update(xlsx_path, mode='missing', target_date=None,
               from_date=None, to_date=None, fetch_online=True):

    print(f"\n{'='*60}")
    print(f"  1:1 Prognosis — Оновлення прогнозу")
    print(f"{'='*60}\n")

    # ── Завантаження Kp ───────────────────────────────────────────────
    kp_data = {}
    if fetch_online:
        print("Завантажую геомагнітні дані...")
        kp_data = fetch_all(verbose=True)
        print(f"Отримано Kp для {len(kp_data)} дат\n")
    else:
        print("(offline режим — Kp з Excel)\n")

    # ── Відкриваємо Excel ─────────────────────────────────────────────
    print(f"Відкриваю: {xlsx_path}")
    backup = str(xlsx_path).replace('.xlsx', '_backup.xlsx')
    shutil.copy2(xlsx_path, backup)
    print(f"Резервна копія: {backup}\n")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['ДАНІ_ЩОДЕННІ']

    updated = 0; skipped = 0; errors = 0

    # ── Обробка рядків ────────────────────────────────────────────────
    for row_idx in range(2, ws.max_row + 1):
        d = parse_date(ws.cell(row_idx, COL_DATE).value)
        if d is None: continue

        # Фільтрація по даті
        if target_date and d != target_date: continue
        if from_date and d < from_date: continue
        if to_date   and d > to_date:   continue

        jy_raw  = ws.cell(row_idx, COL_JY).value or ''
        p_exist = ws.cell(row_idx, COL_P).value or ''

        # Пропускаємо якщо вже є і режим 'missing'
        if mode == 'missing' and p_exist:
            skipped += 1
            continue

        # Отримуємо Kp
        kp = None
        if d in kp_data:
            kp = kp_data[d]['kp']
        else:
            # Беремо з Excel якщо є
            kp_cell = ws.cell(row_idx, COL_KP).value
            if kp_cell:
                try: kp = float(kp_cell)
                except: pass

        if kp is None:
            # Немає Kp — використовуємо середнє значення 3.0
            kp = 3.0

        # Оновлюємо Kp, Storm, Sn, Dst в Excel якщо прийшли онлайн
        if fetch_online and d in kp_data:
            r = kp_data[d]
            ws.cell(row_idx, COL_KP).value    = r['kp']
            ws.cell(row_idx, COL_STORM).value = r['storm']
            if r['sn'] is not None:
                ws.cell(row_idx, COL_SN).value = r['sn']
            if r.get('dst') is not None:
                ws.cell(row_idx, COL_DST).value = r['dst']

        # Контекст затемнення
        eclipse_ctx = detect_eclipse_context(ws, row_idx)

        # Генерація тексту
        try:
            sn  = int(ws.cell(row_idx, COL_SN).value or 0)
            dst = ws.cell(row_idx, COL_DST).value
            dst = float(dst) if dst is not None else (kp_data[d]['dst'] if d in kp_data else None)
            f107 = kp_data[d].get('f107') if d in kp_data else None
            text, sc, lb = format_day(jy_raw, kp, eclipse_context=eclipse_ctx,
                                      sn=sn, dst=dst, f107=f107)
        except Exception as e:
            print(f"  ⚠ {d}: помилка генерації — {e}")
            errors += 1
            continue

        # Записуємо P
        cell_p = ws.cell(row_idx, COL_P)
        cell_p.value = text
        cell_p.alignment = Alignment(wrap_text=True, vertical='top')

        # Колір фону
        color = SCORE_COLORS.get(sc, "DDEBF7")
        cell_p.fill = PatternFill("solid", fgColor=color)

        # Оновлюємо col O (Оцінка дня короткий формат)
        cell_o = ws.cell(row_idx, COL_SCORE)
        cell_o.value = SCORE_LABELS_SHORT.get(sc, '⚪Нейтральний')

        updated += 1
        if updated % 50 == 0:
            print(f"  ... оброблено {updated} рядків")

    # ── Зберігаємо ────────────────────────────────────────────────────
    print(f"\nЗберігаю: {xlsx_path}")
    wb.save(xlsx_path)
    wb.close()

    print(f"\n{'─'*40}")
    print(f"✅ Оновлено:  {updated}")
    print(f"⏭  Пропущено: {skipped}")
    print(f"❌ Помилок:   {errors}")
    print(f"{'─'*40}\n")
    return updated

# ── CLI ───────────────────────────────────────────────────────────────
if __name__ == '__main__':
    p = argparse.ArgumentParser(description='Оновлення прогнозу в Excel')
    p.add_argument('xlsx', nargs='?',
                   default=r'prognoz_2025_2026.xlsx',
                   help='Шлях до Excel файлу')
    p.add_argument('--all', action='store_true',
                   help='Перезаписати всі P (не тільки порожні)')
    p.add_argument('--date', help='Тільки одна дата (YYYY-MM-DD)')
    p.add_argument('--from', dest='from_date', help='Початок діапазону')
    p.add_argument('--to',   dest='to_date',   help='Кінець діапазону')
    p.add_argument('--offline', action='store_true',
                   help='Не завантажувати Kp з інтернету')
    args = p.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        # Шукаємо в стандартних місцях
        candidates = list(Path('.').glob('*prognoz*.xlsx'))
        if candidates:
            xlsx = candidates[0]
            print(f"Знайдено: {xlsx}")
        else:
            print(f"❌ Файл не знайдено: {args.xlsx}")
            sys.exit(1)

    run_update(
        xlsx_path   = xlsx,
        mode        = 'all' if args.all else 'missing',
        target_date = date.fromisoformat(args.date) if args.date else None,
        from_date   = date.fromisoformat(args.from_date) if args.from_date else None,
        to_date     = date.fromisoformat(args.to_date)   if args.to_date   else None,
        fetch_online= not args.offline,
    )
